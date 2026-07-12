from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
import io
import csv
import json
from openpyxl import load_workbook
from datetime import datetime

from ..core.database import get_db
from ..core.deps import get_current_user, get_current_workspace
from ..models.user import User
from ..models.workspace import Workspace
from ..models.import_job import ImportJob as ImportJobModel
from ..models.inventory import InventoryItem as InventoryItemModel
from ..schemas.imports import ImportJob, ImportPreviewResponse
from ..services.activity_service import create_activity

router = APIRouter()

INVENTORY_MAPPING = {
    "sku": ["sku", "stok kodu", "stok_kodu"],
    "name": ["ürün / malzeme adı", "ad", "isim", "ürün adı", "urun adi", "name", "ürün"],
    "category": ["kategori", "category"],
    "unit": ["birim", "unit"],
    "quantity": ["miktar", "stok", "adet", "quantity", "qty"],
    "min_quantity": ["kritik stok", "minimum stok", "min stok", "min_quantity", "kritik_stok"],
    "purchase_price": ["alış fiyatı", "maliyet", "purchase_price", "alis_fiyati"],
    "sale_price": ["satış fiyatı", "fiyat", "sale_price", "satis_fiyati"],
    "supplier": ["tedarikçi", "supplier", "tedarikci"],
    "warehouse_location": ["depo / lokasyon", "konum", "raf", "warehouse_location", "depo", "lokasyon"],
    "notes": ["notlar", "açıklama", "notes", "not"]
}

def find_mapping(header: str):
    header = str(header).lower().strip()
    for field, variations in INVENTORY_MAPPING.items():
        if header in variations:
            return field
    return None

@router.post("/inventory/preview", response_model=ImportPreviewResponse)
async def preview_inventory_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
    user: User = Depends(get_current_user),
) -> Any:
    contents = await file.read()
    filename = file.filename or ""
    normalized_filename = filename.lower()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Dosya boyutu 10 MB sınırını aşıyor.")
    if not normalized_filename.endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=400, detail="Desteklenmeyen dosya formatı. Sadece .csv ve .xlsx desteklenir.")
    
    rows = []
    try:
        if normalized_filename.endswith(".csv"):
            decoded = contents.decode("utf-8").splitlines()
            reader = csv.DictReader(decoded)
            for row in reader:
                rows.append(row)
        elif normalized_filename.endswith(".xlsx"):
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip if entire row is truly empty (all None or whitespace)
                if any(v is not None and str(v).strip() != "" for v in row):
                    rows.append(dict(zip(headers, row)))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Dosya okuma hatası: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="Dosya boş veya veri bulunamadı.")

    valid_rows = []
    invalid_rows = []
    skipped_rows_count = 0

    for idx, raw_row in enumerate(rows):
        mapped_row = {}
        # Clean the row keys and map them
        cleaned_raw_row = {str(k).strip(): v for k, v in raw_row.items() if k is not None}
        
        for k, v in cleaned_raw_row.items():
            field = find_mapping(k)
            if field:
                mapped_row[field] = v
        
        errors = []
        
        # Helper to check if value is effectively empty
        def is_empty(val):
            if val is None: return True
            if str(val).strip() == "": return True
            return False

        # Mandatory fields validation
        if is_empty(mapped_row.get("name")):
            errors.append({"field": "Ürün Adı", "message": "Ürün / Malzeme Adı boş olamaz."})
        
        if is_empty(mapped_row.get("unit")):
            errors.append({"field": "Birim", "message": "Birim alanı zorunludur."})
        
        # Miktar (Quantity) validation - STRICT
        qty_val = mapped_row.get("quantity")
        if is_empty(qty_val):
            errors.append({"field": "Miktar", "message": "Miktar alanı zorunludur."})
        else:
            try:
                # Convert to string first to handle numeric vs string input reliably
                qty_str = str(qty_val).replace(',', '.').strip()
                mapped_row["quantity"] = float(qty_str)
            except (ValueError, TypeError):
                errors.append({"field": "Miktar", "message": "Miktar sayısal olmalıdır."})

        # Numeric validation for other optional fields
        # Kritik Stok
        min_qty_val = mapped_row.get("min_quantity")
        if not is_empty(min_qty_val):
            try:
                mapped_row["min_quantity"] = float(str(min_qty_val).replace(',', '.').strip())
            except:
                errors.append({"field": "Kritik Stok", "message": "Kritik stok sayısal olmalıdır."})
        else:
            mapped_row["min_quantity"] = 0.0

        # Prices
        for price_field, label in [("purchase_price", "Alış Fiyatı"), ("sale_price", "Satış Fiyatı")]:
            val = mapped_row.get(price_field)
            if not is_empty(val):
                try:
                    mapped_row[price_field] = float(str(val).replace(',', '.').strip())
                except:
                    errors.append({"field": label, "message": f"{label} sayısal olmalıdır."})
            else:
                mapped_row[price_field] = 0.0

        if errors:
            invalid_rows.append({"row_number": idx + 2, "errors": errors, "data": cleaned_raw_row})
        else:
            valid_rows.append(mapped_row)

    # Create ImportJob
    import_job = ImportJobModel(
        workspace_id=workspace.id,
        user_id=user.id,
        import_type="inventory",
        filename=filename,
        status="previewed",
        total_rows=len(rows),
        valid_rows=len(valid_rows),
        invalid_rows=len(invalid_rows),
        skipped_rows=skipped_rows_count,
        preview_json=valid_rows,
        error_report_json=invalid_rows
    )
    db.add(import_job)
    db.commit()
    db.refresh(import_job)

    return {
        "import_job_id": import_job.id,
        "total_rows": len(rows),
        "valid_rows": len(valid_rows),
        "invalid_rows": len(invalid_rows),
        "skipped_rows": skipped_rows_count,
        "preview_rows": valid_rows[:10],
        "errors": invalid_rows[:50]
    }

@router.post("/inventory/confirm")
async def confirm_inventory_import(
    import_job_id: int = Form(...),
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
    user: User = Depends(get_current_user),
) -> Any:
    job = db.query(ImportJobModel).filter(
        ImportJobModel.id == import_job_id,
        ImportJobModel.workspace_id == workspace.id
    ).first()
    
    if not job:
        raise HTTPException(status_code=404, detail="Aktarım kaydı bulunamadı.")
    
    if job.status == "completed":
        raise HTTPException(status_code=400, detail="Bu aktarım zaten tamamlanmış.")

    valid_items = job.preview_json or []
    
    imported_count = 0
    skipped_count = 0
    
    for row_data in valid_items:
        try:
            # Check for existing SKU in this workspace
            sku = row_data.get("sku")
            existing_item = None
            if sku:
                existing_item = db.query(InventoryItemModel).filter(
                    InventoryItemModel.workspace_id == workspace.id,
                    InventoryItemModel.sku == sku
                ).first()

            if existing_item:
                # Update existing
                for key, value in row_data.items():
                    setattr(existing_item, key, value)
                existing_item.update_status()
                db.add(existing_item)
            else:
                # Create new
                item = InventoryItemModel(
                    **row_data,
                    workspace_id=workspace.id
                )
                item.update_status()
                db.add(item)
            
            imported_count += 1
        except Exception as e:
            print(f"Error importing row: {str(e)}")
            skipped_count += 1
            continue
    
    job.status = "completed"
    job.imported_rows = imported_count
    job.skipped_rows = skipped_count
    db.commit()
    
    create_activity(
        db, workspace.id, user.id, "inventory", 0, "import",
        f"{imported_count} adet stok kalemi Excel'den aktarıldı."
    )
    
    return {
        "success": True,
        "import_job_id": job.id,
        "status": "completed",
        "imported_rows": imported_count,
        "skipped_rows": skipped_count,
        "message": "Stok listesi başarıyla aktarıldı."
    }

@router.get("/jobs", response_model=List[ImportJob])
def list_import_jobs(
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
) -> Any:
    return db.query(ImportJobModel).filter(ImportJobModel.workspace_id == workspace.id).order_by(ImportJobModel.created_at.desc()).all()

@router.get("/{import_job_id}", response_model=ImportJob)
def get_import_job(
    import_job_id: int,
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
) -> Any:
    job = db.query(ImportJobModel).filter(
        ImportJobModel.id == import_job_id,
        ImportJobModel.workspace_id == workspace.id
    ).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job
