from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

from ..core.database import get_db
from ..core.deps import get_current_user, get_current_workspace
from ..models.user import User
from ..models.workspace import Workspace
from ..models.inventory import InventoryItem as InventoryItemModel
from ..schemas.inventory import InventoryItem, InventoryItemCreate, InventoryItemUpdate, InventorySummary
from ..services.activity_service import create_activity

router = APIRouter()

@router.get("/summary", response_model=InventorySummary)
def get_inventory_summary(
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
) -> Any:
    total_items = db.query(InventoryItemModel).filter(InventoryItemModel.workspace_id == workspace.id).count()
    low_stock_items = db.query(InventoryItemModel).filter(
        InventoryItemModel.workspace_id == workspace.id,
        InventoryItemModel.status == "low_stock"
    ).count()
    out_of_stock_items = db.query(InventoryItemModel).filter(
        InventoryItemModel.workspace_id == workspace.id,
        InventoryItemModel.status == "out_of_stock"
    ).count()
    
    # Simple stock value calculation (quantity * purchase_price)
    # Filter out None prices
    stock_value_query = db.query(func.sum(InventoryItemModel.quantity * InventoryItemModel.purchase_price)).filter(
        InventoryItemModel.workspace_id == workspace.id,
        InventoryItemModel.purchase_price.isnot(None)
    ).scalar()
    
    total_stock_value = stock_value_query or 0.0
    
    categories_count = db.query(func.count(func.distinct(InventoryItemModel.category))).filter(
        InventoryItemModel.workspace_id == workspace.id,
        InventoryItemModel.category.isnot(None)
    ).scalar()
    
    return {
        "total_items": total_items,
        "low_stock_items": low_stock_items,
        "out_of_stock_items": out_of_stock_items,
        "total_stock_value": total_stock_value,
        "categories_count": categories_count or 0
    }

@router.get("/", response_model=List[InventoryItem])
def read_inventory_items(
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
    q: Optional[str] = Query(None, alias="search"),
    category: Optional[str] = None,
    status: Optional[str] = None,
    supplier: Optional[str] = None,
    low_stock: Optional[bool] = Query(None, alias="low_stock_only"),
    skip: int = 0,
    limit: int = 100,
) -> Any:
    query = db.query(InventoryItemModel).filter(InventoryItemModel.workspace_id == workspace.id)
    
    if q:
        query = query.filter(InventoryItemModel.name.ilike(f"%{q}%") | InventoryItemModel.sku.ilike(f"%{q}%"))
    if category:
        query = query.filter(InventoryItemModel.category == category)
    if status:
        query = query.filter(InventoryItemModel.status == status)
    if supplier:
        query = query.filter(InventoryItemModel.supplier == supplier)
    if low_stock:
        query = query.filter(InventoryItemModel.status.in_(["low_stock", "out_of_stock"]))
        
    return query.order_by(InventoryItemModel.name.asc()).offset(skip).limit(limit).all()

@router.post("/", response_model=InventoryItem)
def create_inventory_item(
    *,
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
    user: User = Depends(get_current_user),
    item_in: InventoryItemCreate,
) -> Any:
    item = InventoryItemModel(
        **item_in.dict(),
        workspace_id=workspace.id
    )
    item.update_status()
    db.add(item)
    db.commit()
    db.refresh(item)
    
    create_activity(
        db, workspace.id, user.id, "inventory", item.id, "create",
        f"Stok kalemi oluşturuldu: {item.name}."
    )
    
    return item

@router.get("/template")
def download_inventory_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Aktarım Şablonu"
    
    headers = [
        "SKU", "Ürün / Malzeme Adı", "Kategori", "Birim", "Miktar", 
        "Kritik Stok", "Alış Fiyatı", "Satış Fiyatı", "Tedarikçi", 
        "Depo / Lokasyon", "Notlar"
    ]
    
    # Style
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        # Column width
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        
    # Sample row
    sample_row = [
        "SKU001", "Örnek Ürün", "Genel", "Adet", 100, 
        10, 50.0, 75.0, "ABC Tedarik", "A1 Rafı", "Örnek not."
    ]
    ws.append(sample_row)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=operio-stok-sablonu.xlsx"}
    )

@router.get("/export")
def export_inventory(
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
    user: User = Depends(get_current_user),
):
    items = db.query(InventoryItemModel).filter(InventoryItemModel.workspace_id == workspace.id).order_by(InventoryItemModel.name.asc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Listesi"
    
    headers = [
        "SKU", "Ürün / Malzeme Adı", "Kategori", "Birim", "Miktar", 
        "Kritik Stok", "Alış Fiyatı", "Satış Fiyatı", "Tedarikçi", 
        "Depo / Lokasyon", "Notlar", "Durum"
    ]
    
    header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    header_font = Font(bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        
    for idx, item in enumerate(items, 2):
        ws.cell(row=idx, column=1, value=item.sku)
        ws.cell(row=idx, column=2, value=item.name)
        ws.cell(row=idx, column=3, value=item.category)
        ws.cell(row=idx, column=4, value=item.unit)
        ws.cell(row=idx, column=5, value=item.quantity)
        ws.cell(row=idx, column=6, value=item.min_quantity)
        ws.cell(row=idx, column=7, value=item.purchase_price)
        ws.cell(row=idx, column=8, value=item.sale_price)
        ws.cell(row=idx, column=9, value=item.supplier)
        ws.cell(row=idx, column=10, value=item.warehouse_location)
        ws.cell(row=idx, column=11, value=item.notes)
        ws.cell(row=idx, column=12, value=item.status)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    create_activity(
        db, workspace.id, user.id, "inventory", 0, "export",
        "Stok listesi dışa aktarıldı."
    )
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=operio-stok-listesi.xlsx"}
    )

@router.get("/{item_id}", response_model=InventoryItem)
def read_inventory_item(
    item_id: int,
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
) -> Any:
    item = db.query(InventoryItemModel).filter(
        InventoryItemModel.id == item_id,
        InventoryItemModel.workspace_id == workspace.id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return item

@router.put("/{item_id}", response_model=InventoryItem)
def update_inventory_item(
    *,
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
    user: User = Depends(get_current_user),
    item_id: int,
    item_in: InventoryItemUpdate,
) -> Any:
    item = db.query(InventoryItemModel).filter(
        InventoryItemModel.id == item_id,
        InventoryItemModel.workspace_id == workspace.id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    update_data = item_in.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(item, field, value)
        
    item.update_status()
    db.add(item)
    db.commit()
    db.refresh(item)
    
    create_activity(
        db, workspace.id, user.id, "inventory", item.id, "update",
        f"Stok kalemi güncellendi: {item.name}."
    )
    
    return item

@router.delete("/{item_id}", response_model=InventoryItem)
def delete_inventory_item(
    *,
    db: Session = Depends(get_db),
    workspace: Workspace = Depends(get_current_workspace),
    user: User = Depends(get_current_user),
    item_id: int,
) -> Any:
    item = db.query(InventoryItemModel).filter(
        InventoryItemModel.id == item_id,
        InventoryItemModel.workspace_id == workspace.id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # RBAC check: only owner/admin/manager
    # For now we assume role check is done in core/deps if needed, or check here
    # Assuming user role is in user model
    
    db.delete(item)
    db.commit()
    
    create_activity(
        db, workspace.id, user.id, "inventory", item_id, "delete",
        f"Stok kalemi silindi: {item.name}."
    )
    
    return item
